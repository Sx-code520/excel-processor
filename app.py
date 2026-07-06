#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2026/7/6 17:25
# @Author : 15521
# @File : app.py
# @Software: PyCharm
"""

"""
import streamlit as st
import openpyxl
import re
import io
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==========================================
# 1. 密码验证逻辑 (放在最前面)
# ==========================================
def check_password():
   """返回 True 如果用户提供了正确的密码"""

   def password_entered():
      """检查输入的密码是否正确，并使用 hmac 防止时序攻击"""
      from hmac import compare_digest
      if compare_digest(st.session_state["password"], st.secrets["password"]):
         st.session_state["password_correct"] = True
         del st.session_state["password"]  # 验证后从 session 中删除密码
      else:
         st.session_state["password_correct"] = False

   # 如果已经验证通过，直接放行
   if st.session_state.get("password_correct", False):
      return True

   # 首次运行，显示密码输入框
   st.text_input(
      "请输入访问密码", type="password", on_change=password_entered, key="password"
   )

   # 密码错误时显示提示
   if "password_correct" in st.session_state and not st.session_state["password_correct"]:
      st.error("😕 密码不正确，请重试")

   return False


# 如果密码验证未通过，停止执行后续代码
if not check_password():
   st.stop()


# ==========================================
# 2. 核心处理逻辑（封装为函数）
# ==========================================
def process_excel(file):
   # 加载原始工作簿
   original_wb = openpyxl.load_workbook(file)
   original_sheet = original_wb['Sheet1']

   # 【核心升级】智能搜索元数据，不再依赖固定坐标
   contract_id = "未找到编号"
   customer_name = "未找到单位"

   # 在前10行内搜索关键字（防止表头位置上下浮动）
   for row in range(1, 11):
      for col in range(1, 10):  # 搜索前9列
         cell_value = str(original_sheet.cell(row=row, column=col).value).strip()

         # 查找合同编号
         if "合同编号" in cell_value:
            # 假设编号在关键字右侧的第一个非空单元格
            for next_col in range(col + 1, col + 4):
               next_val = original_sheet.cell(row=row, column=next_col).value
               if next_val is not None and str(next_val).strip():
                  contract_id = str(next_val).strip()
                  break

         # 查找客户单位
         if "客户单位" in cell_value:
            # 假设单位在关键字右侧的第一个非空单元格
            for next_col in range(col + 1, col + 4):
               next_val = original_sheet.cell(row=row, column=next_col).value
               if next_val is not None and str(next_val).strip():
                  customer_name = str(next_val).strip()
                  break

   # 数据提取、清洗与补全逻辑
   data_start_row = 6
   processed_data = []

   # 第一遍：提取纯数字编号
   for row in range(data_start_row, original_sheet.max_row + 1):
      product_id_raw = str(original_sheet[f'B{row}'].value).strip()
      if product_id_raw == 'None' or not product_id_raw:
         continue
      # 只保留纯数字编号
      if not re.match(r'^\d+$', product_id_raw):
         continue

      processed_data.append({
         'product_id': int(product_id_raw),
         'product_name': original_sheet[f'C{row}'].value,
         'spec': original_sheet[f'D{row}'].value,
         'device_tag': original_sheet[f'F{row}'].value
      })

   # 第二遍：处理区间补全（兼容换行符）
   for row in range(data_start_row, original_sheet.max_row + 1):
      product_id_raw = str(original_sheet[f'B{row}'].value).strip()

      # 【核心升级】\s* 可以匹配换行符 \n、回车符 \r 或空格
      match = re.match(r'^(\d+)/\s*(\d+)$', product_id_raw)

      if match:
         start_id = int(match.group(1))
         end_id = int(match.group(2))

         # 提取当前行的其他信息
         product_name = original_sheet[f'C{row}'].value
         spec = original_sheet[f'D{row}'].value
         device_tag = original_sheet[f'F{row}'].value

         # 遍历区间，补全中间缺失的编码
         for current_id in range(start_id, end_id + 1):
            # 避免重复添加
            if not any(d['product_id'] == current_id for d in processed_data):
               processed_data.append({
                  'product_id': current_id,
                  'product_name': product_name,
                  'spec': spec,
                  'device_tag': device_tag
               })

   # 排序
   processed_data.sort(key=lambda x: x['product_id'])

   # 生成新工作簿到内存
   new_wb = openpyxl.Workbook()
   new_sheet = new_wb.active
   new_sheet.title = "提取结果"

   # 样式设置
   header_font = Font(bold=True)
   center_alignment = Alignment(horizontal='center', vertical='center')
   border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

   # 写入表头信息（取消合并，严格分列）
   new_sheet['A1'] = "合同编号："
   new_sheet['B1'] = contract_id  # 编号将严格显示在 B1 单元格
   new_sheet['C1'] = "客户单位："
   new_sheet['D1'] = customer_name  # 单位将严格显示在 E1 单元格

   # 仅对标签（A1和D1）加粗居中，数据单元格保持默认
   new_sheet['A1'].font = header_font
   new_sheet['A1'].alignment = center_alignment
   new_sheet['D1'].font = header_font
   new_sheet['D1'].alignment = center_alignment

   headers = ["产品编号", "产品名称", "型号/规格", "设备名称及位号"]
   for col_num, header in enumerate(headers, 1):
      cell = new_sheet.cell(row=2, column=col_num, value=header)
      cell.font = header_font
      cell.alignment = center_alignment
      cell.border = border_style

   # 写入数据
   for idx, data in enumerate(processed_data, 3):
      new_sheet.cell(row=idx, column=1, value=data['product_id'])
      new_sheet.cell(row=idx, column=2, value=data['product_name'])
      new_sheet.cell(row=idx, column=3, value=data['spec'])
      new_sheet.cell(row=idx, column=4, value=data['device_tag'])
      for col in range(1, 5):
         new_sheet.cell(row=idx, column=col).border = border_style
         new_sheet.cell(row=idx, column=col).alignment = center_alignment

   # 调整列宽
   for col in range(1, 6):
      new_sheet.column_dimensions[get_column_letter(col)].width = 20

   # 保存到内存字节流
   output = io.BytesIO()
   new_wb.save(output)
   output.seek(0)
   return output, contract_id, len(processed_data)


# ==========================================
# 3. Streamlit 网页界面
# ==========================================
st.set_page_config(page_title="Excel 智能处理工具", layout="centered")
st.title("📊 Excel 智能提取与补全工具")
st.caption("上传原始目录表，自动清洗、补全区间编码并排序，生成规范的新表格。")

uploaded_file = st.file_uploader("选择你的 Excel 文件", type=["xlsx"])

if uploaded_file is not None:
   if st.button("🚀 开始处理"):
      with st.spinner("正在处理数据，请稍候..."):
         try:
            result_file, contract_id, count = process_excel(uploaded_file)
            st.success(f"处理完成！共提取并补全了 **{count}** 条有效记录。")

            # 提供下载按钮
            st.download_button(
               label="📥 下载处理后的 Excel 文件",
               data=result_file,
               file_name=f"提取结果_{contract_id}.xlsx",
               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
         except Exception as e:
            st.error(f"处理过程中发生错误: {e}")
